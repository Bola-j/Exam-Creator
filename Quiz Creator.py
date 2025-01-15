from tkinter import *
from tkinter import messagebox as mb
import openpyxl.workbook
from ttkbootstrap.constants import *
import ttkbootstrap as tb
import openpyxl

# Global dictionaries to store questions and correct answers.
questions = {}
correct_answers = {}


def create_quiz(Quiz): # Create the quiz excel sheet
    try:
        # Load existing quizzes file and create a new sheet for the quiz having the new quiz's name.
        file = openpyxl.load_workbook("quizzes.xlsx")
        file.create_sheet(Quiz, index=0)
        file.save("quizzes.xlsx")
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist.
        file = openpyxl.Workbook()
        file.create_sheet(Quiz, index=0)
        file.save("quizzes.xlsx")
    set_fields()

def set_fields(): # Function to set field headers for a new quiz sheet.
    file = openpyxl.load_workbook("quizzes.xlsx")
    file.active["A1"] = "type"
    file.active["B1"] = "ques"
    file.active["C1"] = "ch1"
    file.active["D1"] = "ch2"
    file.active["E1"] = "ch3"
    file.active["F1"] = "ch4"
    file.active["G1"] = "true"
    file.active["H1"] = "weight"
    file.save("quizzes.xlsx")


# Initialize the main application window.
homescreen = tb.Window()
homescreen.title("Quiz Creator")
homescreen.geometry("500x200")


# Define styles for buttons.
stylesuccess = tb.Style()
stylesuccess.configure("success.TButton", font=("Helvetica", 70))
style_quiz_taking = tb.Style()
style_quiz_taking.configure("primary.TButton", font=("Helvetica", 10))


# Main screen label.
application_title = tb.Label(homescreen, text="Welcome to Nu Quiz Maker (NQM)", font=("Arial", 20))

# Function for student interface to take quizzes.
def end_student():
    #calling global variables to edit and use
    global current_page, exams_per_page, total_mark, questions, correct_answers
    questions = {}
    correct_answers = {}
    try:    
        exams = openpyxl.load_workbook("quizzes.xlsx").sheetnames
    except FileNotFoundError:
             # Show message if no quizzes are available.(File doesn't exist)
            mb.showinfo(title="No Quizzes", message="No quizzes are available")
            return
    current_page = 0
    exams_per_page = 7
    total_mark = 0
    # Function to open a quiz and display its questions.
    def exam_page(sheet_name):
        global collected_marks, radio_buttons, selected_answers, question_index, answer
        # Load quiz sheet and extract questions and answers.
        quiz = openpyxl.load_workbook("quizzes.xlsx")[sheet_name]
        num_of_questions = quiz.max_row
        questions = {}
        correct_answers = {}
        #retrieving questions and answers from excel sheet
        for x in range(2, num_of_questions+1):
            if quiz[f"A{x}"].value == "mcq":
                questions[quiz[f"B{x}"].value] = [quiz[f"C{x}"].value, quiz[f"D{x}"].value, quiz[f"E{x}"].value, quiz[f"F{x}"].value]
            elif quiz[f"A{x}"].value == "comp":
                questions[quiz[f"B{x}"].value] = quiz[f"G{x}"].value
            else:
                break
            correct_answers[quiz[f"B{x}"].value] = (quiz[f"G{x}"].value, quiz[f"H{x}"].value)    
        # Create a new window for the quiz.
        quiz_window = Toplevel()
        quiz_window.title("Quiz")
        quiz_window.geometry("700x700")

        radio_buttons = []  # Store radio buttons for multiple choice questions.
        selected_answers = ["no_answer"] * len(questions)  # Store selected answers.
        answer = StringVar()  # Variable for the selected answer.
        question_index = 0  # Current question index.
        collected_marks = 0
        
        def display_question(): # Function to display the current question.
            global radio_buttons
            
            if question_index == len(questions)-1: # change text of 'next' button to 'submit' when final question is reached
                next_button.config(text="Submit")

            
            #clear and delete radiobuttons and complete answer entry before showing the next questions
            for i in radio_buttons:
                i.destroy()
            radio_buttons.clear()
            comp_ans_field.delete(0, END)
            comp_ans_field.forget()

            question_label.config(text= list(questions.keys())[question_index]) # update question label
            #deactivate back button on 1st question
            deactivate_prev_button()
            if type(questions[list(questions.keys())[question_index]]) is list: # displaying question options when it is mcq
                current_options = questions[list(questions.keys())[question_index]]
                answer.set(selected_answers[question_index])  # saves picked answer
                # sets radiobuttons
                for option in (current_options):
                    radiobutton = Radiobutton(quiz_window, font=("Arial", 12), text=option, variable=(answer), value=(option))
                    radiobutton.pack(anchor="w")
                    radio_buttons.append(radiobutton)
            else: # displaying question answer entry when it is complete
                comp_ans_field.pack(pady= 10)
        
        def retrieve_answer():
            # To retrieve and save answer
            global selected_answers
            if type(questions[list(questions.keys())[question_index]]) is not tuple: #MCQ
                selected_answers[question_index] = answer.get()
            else: #complete
                selected_answers[question_index] = comp_ans_field.get()
        
        def next_question():
        # function for next question
            global question_index, total_mark, collected_marks
            retrieve_answer()
            if question_index < len(questions)-1:
                question_index += 1
                display_question()
            else:
                # grade your answers and display result
                for i in range(len(questions)):
                    if selected_answers[i].lower() == list(correct_answers.values())[i][0].lower():
                        collected_marks += list(correct_answers.values())[i][1]
                    total_mark += list(correct_answers.values())[i][1]
                #message box with the quiz grade
                mb.showinfo(title='Grade',message=f'You got {collected_marks}/{total_mark}\n%{(collected_marks/total_mark)*100}')
                collected_marks = 0
                total_mark = 0
                quiz_window.destroy()
        
        def previous_question():
        # function to go back
            global question_index
            next_button.config(text="Next") # change text from 'submit' to 'next' when going back
            retrieve_answer()
            if question_index !=0:
                question_index -= 1
                display_question()
        
        #deactivate back button on 1st question
        def deactivate_prev_button():
            if question_index == 0:
                prev_button["state"] = DISABLED
            else:
                prev_button["state"] = NORMAL

         # Create question label and input field.
        question_label = Label(quiz_window, font=("Arial", 18), pady=20, padx=10)
        question_label.pack()
        comp_ans_field = Entry( quiz_window ,font= ("Helvetica", 40))

        # button for previous question
        prev_button = Button(quiz_window, text="Back", command=previous_question)
        prev_button.pack(side=LEFT, padx=10)

        # button for next question
        next_button = Button(quiz_window, text="Next", command=next_question)
        next_button.pack(side=RIGHT, padx=10)

        #to display first question
        display_question()

        quiz_window.mainloop()



    # Function to switch from 1 set of exams to another
    def update_display():
        # Delete previous exam buttons
        for i in exam_display.winfo_children():
            i.destroy()

        # Exam set for current page
        start = current_page * exams_per_page
        current_exams = exams[start:start+exams_per_page]

        # Display exams as buttons
        for exam in (current_exams):
            if not(exam.startswith("Sheet")):
                button = Button(exam_display, text=f"{exam}", command=lambda e=exam: exam_page(e))
                button.pack(pady=10)

        # Update button state of buttons
        if current_page == 0:
            prev_button["state"] = DISABLED
        else:
            prev_button["state"] = NORMAL
        
        if (current_page + 1) * exams_per_page >= len(exams)-1:
            next_button["state"] = DISABLED
        else:
            next_button["state"] = NORMAL

    # Previous button function
    def prev_page():
        global current_page
        if current_page > 0:
            current_page -= 1
            update_display()

    # Next button function
    def next_page():
        global current_page
        if (current_page + 1) * exams_per_page < len(exams):
            current_page += 1
            update_display()

    # Create quizzes page.
    quizzes_page = Toplevel()
    quizzes_page.title("Quizzes")

    exam_display = Frame(quizzes_page)
    exam_display.pack(pady=20)

    prev_button = Button(quizzes_page, text="<- Previous", command=prev_page)
    prev_button.pack(side=LEFT, padx=10)
    next_button = Button(quizzes_page, text="Next ->", command=next_page)
    next_button.pack(side=RIGHT, padx=10)

    update_display()
    quizzes_page.geometry("500x500")
    quizzes_page.mainloop()

#tutor  end
def end_tutor():
    global questions, correct_answers, correct_answer
    questmaker = Toplevel()
    tabs = tb.Notebook(questmaker, width=1750, height=1000)
    tabs.pack(pady=20)
    questmaker.title("Create questions")
    questmaker.geometry("1000x950")

    correct_answer= IntVar()

    def save_ques(): # saving questions, answers and wrights for mcq questions in questions and correct_answers
        global questions, correct_answers 
        questions[mcq_question.get()] = [ans1.get(), ans2.get() ,ans3.get() ,ans4.get()] #saving MCQ and its choices in dictionary
        try:
            correct_answers[mcq_question.get()] = (list(questions[mcq_question.get()])[correct_answer.get()-1] , float(weight_mcq.get()))
        except: # if question weight is entered inproperly, will equal 1 as a default value
            correct_answers[mcq_question.get()] = (list(questions[mcq_question.get()])[correct_answer.get()-1] , 1 )

        #clearing enteries and correct_answer
        mcq_question.delete(0, END)
        ans1.delete(0, END)
        ans2.delete(0, END)
        ans3.delete(0, END)
        ans4.delete(0, END)
        weight_mcq.delete(0, END)
        correct_answer.set(None)

    def store_quiz(name): #storing quiz's questions in its sheet
        if (len(questions) != 0) or (len(correct_answers) != 0): #only stores when it contains questions
            create_quiz(name) #create the quiz sheet
            file = openpyxl.load_workbook("quizzes.xlsx") #loading it
            sheet = file.active # working on the quiz sheet (default on on index 0)
            row_index = 2 # row counter
            for x in range(len(questions)):
                choice_field_code = ord("C") #column counter for changing column on storing each mcq option
                sheet[f"B{row_index}"] = list(questions.keys())[x] # storing ques text
                if type(questions[list(questions.keys())[x]]) is tuple: # complete condition
                    sheet[f"A{row_index}"] = "comp" # storing ques type
                    sheet[f"G{row_index}"] = questions[list(questions.keys())[x]][0] # storing correct answer
                    sheet[f"H{row_index}"] = questions[list(questions.keys())[x]][1] # storing ques weight
                else: #mcq condition
                    sheet[f"A{row_index}"] = "mcq" # storing ques type
                    for ans in questions[list(questions.keys())[x]]: # storing ques options
                        sheet[f"{chr(choice_field_code)}{row_index}"] = ans 
                        choice_field_code += 1 #changing column
                    sheet[f"G{row_index}"] = correct_answers[list(questions.keys())[x]][0] # storing correct answer
                    sheet[f"H{row_index}"] = correct_answers[list(questions.keys())[x]][1] # storing ques weight
                row_index += 1 #changing row
            file.save("quizzes.xlsx")
        else: # when it's an empty quiz, shows this message
            mb.showinfo(title="Blank Quiz", message="No questions added in the quiz, it will not be saved")
        questmaker.destroy()
    def savecomp():# saving questions, answers and wrights for complete questions in questions and correct_answers
        global questions, correct_answers
        try:
            questions[comp_question.get()] = ( comp_ans.get() , float(weight_comp.get()))
            correct_answers[comp_question.get()] = ( comp_ans.get() , float(weight_comp.get()))
        except: # if question weight is entered inproperly, will equal 1 as a default value
            questions[comp_question.get()] = ( comp_ans.get() , 1)
            correct_answers[comp_question.get()] = ( comp_ans.get() , 1)
        #clearing enteries and correct_answer
        comp_question.delete(0, END)
        comp_ans.delete(0, END)
        weight_comp.delete(0, END)
# This is the first tab or frame in the quiz creation part, the mcq
    MCQ = tb.Frame(tabs)

    # text and entry box for the name
    quiz_name_label = tb.Label(MCQ, text="Quiz Name", font=("Helvetica", 28))
    quiz_name = tb.Entry(MCQ, font=("Helvetica", 40))

    # this is the text showing that you add the MCQ question and it's entry
    mcq_tabtitle = tb.Label(
        MCQ, text="Multiple Choice Question", font=("Helvetica", 25)
    )
    mcq_question = tb.Entry(MCQ, font=("Helvetica", 40))

    # these are the four entry boxes for the answers
    ans1 = tb.Entry(MCQ, font=("Helvetica", 20))
    ans2 = tb.Entry(MCQ, font=("Helvetica", 20))
    ans3 = tb.Entry(MCQ, font=("Helvetica", 20))
    ans4 = tb.Entry(MCQ, font=("Helvetica", 20))

    # these are the buttons that allow user to pick correct answer
    cor1 = tb.Radiobutton(MCQ, text="Correct", variable=correct_answer, value=1)
    cor2 = tb.Radiobutton(MCQ, text="Correct", variable=correct_answer, value=2)
    cor3 = tb.Radiobutton(MCQ, text="Correct", variable=correct_answer, value=3)
    cor4 = tb.Radiobutton(MCQ, text="Correct", variable=correct_answer, value=4)

    # this is the text and entry for the weight system (if invalid entry sets as 1)
    weight_mcq_label = tb.Label(MCQ, text="question weight", font=("Helvetica", 20))
    weight_mcq = tb.Entry(MCQ, font=("Helvetica", 15))

    # these is the save button(appends MCQ question to the excel file)
    save_but = tb.Button(MCQ, text="Save", command=save_ques)

    # this is the finish button(saves the excel file )
    home_but = tb.Button(
        MCQ, text="Finish", command=lambda: store_quiz(quiz_name.get())
    )

    # here we add to the main tab that the name of the MCQ frame or tab is called "Multiple choice"
    tabs.add(MCQ, text="Multiple choice")

    # like the MCQ tab this is the tab but for the complete questions which creates a tab
    COMP = tb.Frame(tabs)

    # Main text and entry to add the question itself
    comp_tabtitle = tb.Label(COMP, text="Complete Question", font=("Helvetica", 38))
    comp_question = tb.Entry(COMP, font=("Helvetica", 40))

    # This it the text and answer for the complete question
    comp_ans_label = tb.Label(
        COMP, text="Answer for the question", font=("Helvetica", 38)
    )
    comp_ans = tb.Entry(COMP, font=("Helvetica", 40))

    # save button
    save_comp_but = tb.Button(COMP, text="Save", command=savecomp)

    # weight text and entry
    weight_comp = tb.Entry(COMP, font=("Helvetica", 15))
    weight_comp_label = tb.Label(COMP, text="question weight", font=("Helvetica", 20))

    # names the COMP tab or frame
    tabs.add(COMP, text="Complete")

    # this is the packing for all the widgets used (IT HAS TO BE IN ORDER)
    quiz_name_label.pack(pady=0)
    quiz_name.pack(padx=10)
    mcq_tabtitle.pack(pady=0)
    mcq_question.pack(padx=10)

    ans1.pack(pady=5)
    cor1.pack(pady=3)

    ans2.pack(pady=3)
    cor2.pack(pady=3)

    ans3.pack(pady=3)
    cor3.pack(pady=3)

    ans4.pack(pady=3)
    cor4.pack(pady=0)

    weight_mcq_label.pack(pady=0)
    weight_mcq.pack(pady=0)

    save_but.pack(pady=4)
    home_but.pack(pady=0)

    comp_tabtitle.pack(pady=20)
    comp_question.pack(pady=10, padx=10)
    comp_ans_label.pack(pady=10)
    comp_ans.pack(pady=10, padx=10)
    weight_comp_label.pack(pady=0)
    weight_comp.pack(pady=0)
    save_comp_but.pack(pady=20)


# Main 2 buttons for student entry and tutor entry
student = tb.Button(
    homescreen, text="Student", style="success.TButton", width=10, command=end_student
)
tutor = tb.Button(
    homescreen, text="Tutor", style="success.TButton", width=10, command=end_tutor
)

# this shows the button in a grid style to make the 2 buttons appear next to each other
application_title.grid(row=0, column=0, columnspan=2)
student.grid(row=1, column=0)
tutor.grid(row=1, column=1)

# these are some configs for the grid and overall geometry for the main screen
homescreen.attributes("-fullscreen", True)
homescreen.grid_rowconfigure(0, weight=1)
homescreen.grid_rowconfigure(1, weight=3)
homescreen.grid_columnconfigure(0, weight=1)
homescreen.grid_columnconfigure(1, weight=1)

# mainloop is to show that all the functions are working in this screen
homescreen.mainloop()